"""
Excel converter for DocBridge
"""
from typing import Optional, Union, List
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .base import BaseConverter
from ..styles.default import DefaultStyle
from ..formatters.table import TableFormatter
from ..utils import parse_markdown_table


class ExcelConverter(BaseConverter):
    """
    Convert Markdown tables to Excel spreadsheets.
    """

    def __init__(self, style: Optional[DefaultStyle] = None):
        super().__init__(style)
        self.table_formatter = TableFormatter()

    def convert(self, markdown_text: str) -> Workbook:
        """
        Convert markdown text to Excel workbook.

        Args:
            markdown_text: The markdown text to convert

        Returns:
            openpyxl Workbook object
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Extract and convert tables
        tables = self._extract_tables(markdown_text)

        if tables:
            # Write first table to first sheet
            self._write_table_to_sheet(ws, tables[0])

            # Additional tables go to new sheets
            for i, table in enumerate(tables[1:], 2):
                ws_new = wb.create_sheet(title=f"Sheet{i}")
                self._write_table_to_sheet(ws_new, table)
        else:
            # No tables found, write as plain text
            ws['A1'] = markdown_text

        self._document = wb
        return wb

    def convert_file(self, input_path: Union[str, Path],
                     output_path: Optional[Union[str, Path]] = None) -> str:
        """
        Convert markdown file to Excel file.

        Args:
            input_path: Path to input markdown file
            output_path: Path for output file (optional)

        Returns:
            Path to output file
        """
        # Read input file
        markdown_text = self._read_file(input_path)

        # Convert to workbook
        wb = self.convert(markdown_text)

        # Determine output path
        output_path = self._ensure_output_path(input_path, output_path, '.xlsx')

        # Save workbook
        wb.save(output_path)

        return str(output_path)

    def _extract_tables(self, text: str) -> List[dict]:
        """Extract all tables from markdown text."""
        lines = text.split('\n')
        tables = []
        current_table = []

        for line in lines:
            if self.table_formatter.is_table_line(line):
                current_table.append(line)
            else:
                if current_table:
                    table_text = '\n'.join(current_table)
                    table_data = self.table_formatter.parse_table(table_text)
                    if table_data['headers']:
                        tables.append(table_data)
                    current_table = []

        # Don't forget last table
        if current_table:
            table_text = '\n'.join(current_table)
            table_data = self.table_formatter.parse_table(table_text)
            if table_data['headers']:
                tables.append(table_data)

        return tables

    def _write_table_to_sheet(self, ws, table_data: dict) -> None:
        """Write table data to worksheet."""
        headers = table_data['headers']
        rows = table_data['rows']
        alignments = table_data.get('alignments', ['left'] * len(headers))

        # Define styles
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col, cell_value in enumerate(row_data, 1):
                if col <= len(headers):
                    cell = ws.cell(row=row_idx, column=col, value=cell_value)
                    cell.border = thin_border

                    # Apply alignment
                    align = alignments[col - 1] if col <= len(alignments) else 'left'
                    cell.alignment = Alignment(horizontal=align, vertical='center')

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)

            for row in range(1, len(rows) + 2):
                cell = ws[f"{column}{row}"]
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Set row height for header
        ws.row_dimensions[1].height = 25
